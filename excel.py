from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os

BILLS_FILENAME = "bills/all_bills.xlsx"

def save_bill_to_excel(bill_items, total):
    os.makedirs("bills", exist_ok=True)

    # Load or create workbook
    if os.path.exists(BILLS_FILENAME):
        wb = load_workbook(BILLS_FILENAME)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Find the next empty row to start new bill
    start_row = ws.max_row + 2  # 1 blank row between bills

    # Token or Customer Number (incremented)
    customer_number = (ws.max_row // 20) + 1  # crude approx, or you can store separately

    # Write Bill Header
    ws.cell(row=start_row, column=1, value=f"Customer / Token No: {customer_number}").font = bold_font
    ws.cell(row=start_row, column=3, value=f"Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = bold_font

    # Table headers for items
    header_row = start_row + 2
    headers = ["Item", "Quantity", "Price", "Total"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Write each bill item
    for i, item in enumerate(bill_items, start=1):
        row = header_row + i
        ws.cell(row=row, column=1, value=item["name"]).border = thin_border
        ws.cell(row=row, column=2, value=item["qty"]).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=item["price"]).alignment = center_align
        ws.cell(row=row, column=3).border = thin_border
        total_price = item["qty"] * item["price"]
        ws.cell(row=row, column=4, value=total_price).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border

    # Write Total row
    total_row = header_row + len(bill_items) + 1
    ws.cell(row=total_row, column=3, value="Total").font = bold_font
    ws.cell(row=total_row, column=3).alignment = center_align
    ws.cell(row=total_row, column=3).border = thin_border

    ws.cell(row=total_row, column=4, value=total).font = bold_font
    ws.cell(row=total_row, column=4).alignment = center_align
    ws.cell(row=total_row, column=4).border = thin_border

    wb.save(BILLS_FILENAME)
    print(f"✅ Bill appended to {BILLS_FILENAME}")

    # Optional: update daily summary
    save_summary_log(datetime.now().date(), total)


def save_summary_log(date, total):
    date_str = date.strftime("%Y-%m-%d")
    summary_file = f"bills/summary_{date_str}.xlsx"

    if os.path.exists(summary_file):
        wb = load_workbook(summary_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Date", "Total Sales"])

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([now_str, total])
    wb.save(summary_file)
    print(f"📊 Daily summary updated: {summary_file}")
